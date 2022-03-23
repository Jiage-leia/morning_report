from itertools import zip_longest
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from global_variables import *


THIN = Side(border_style="thin", color="000000")

def insert_pa_np_row(sheet_obj, provider, iov_tuple):
    if len(iov_tuple) == 0:
        pass
    pass

def insert_task_row(sheet_obj, panp_iov_dict, task_dict):
    if sum([len(panp_iov_dict[n]) for n in panp_iov_dict]) == 0:
        for key in PA_NP_TASK_LOCATION:
            #put data on the row of the first name of PA_NP
            _row = find_row_per_keyword(sheet_obj, PANP_ANCHOR)
            assert _row != 0
            if key in PA_NP_TASK:
                sheet_obj.cell(row=_row+1, column=column_index_from_string(PA_NP_TASK_LOCATION[key]), value = len(task_dict[key]))
            sheet_obj = merge_md_units(sheet_obj, PA_NP_TASK_LOCATION[key], _row+1, _row+len(PA_NP)+1)
        return sheet_obj
    return sheet_obj

def find_row_per_keyword(sheet_obj, pattern):
    for i, u in enumerate(sheet_obj['A'], 1):
        try:
            if u.value == pattern:
                return i
        except:
            pass
    return 0
        

def insert_md_row(sheet_obj, md, iov_tuple, talk_tuple):
    
    _start_idx, _end_idx = 0, 0
    for i, u in enumerate(sheet_obj['A'], 1):
        try:
            if u.value.upper() == md:
                _start_idx = i
                _end_idx = i
                _inside_counter = 0
                for _iov, _talk in zip_longest(iov_tuple, talk_tuple):
                    # print(_inside_counter)
                    # print([_iov, _talk])
                    _iov_out, _talk_out = [None, None, None], [None, None, None]
                    if _iov:
                        _iov_out = _iov
                    if _talk:
                        _talk_out = _talk
                    
                    _new_row = [None]+_iov_out+[None, None, None]+_talk_out
                    _fill_index = i
                    # print(_new_row)

                    if _inside_counter != 0:
                        _fill_index += _inside_counter
                        sheet_obj.insert_rows(_fill_index)
                    
                    _inside_counter += 1
                    for idx in range(1, len(_new_row)):
                        _cell = sheet_obj.cell(row=_fill_index, column=idx+1)
                        _cell.value = _new_row[idx]
                        _cell.border =  Border(top=THIN, left=THIN, right=THIN, bottom=THIN)
                        # _ = sheet_obj.cell(row=_fill_index, column=idx+1, value=_new_row[idx])
                if _inside_counter > 0:
                    sheet_obj= merge_md_units(sheet_obj, COL_LIM, _start_idx, _start_idx+_inside_counter)
                    _end_idx += _inside_counter
                else:
                    for idx in range(1, 10):
                        _cell = sheet_obj.cell(row=i, column=idx+1)
                        _cell.border =  Border(top=THIN, left=THIN, right=THIN, bottom=THIN)
        except Exception as et:
            # print(et)
            continue
    # sheet_obj = unmerge_cells(sheet_obj, _end_idx)
    return sheet_obj



def merge_md_units(sheet_obj, column, start_idx, end_idx):
    sheet_obj.merge_cells('%s%s:%s%s' % (column, start_idx, column, end_idx-1))
    return sheet_obj

def unmerge_cells(sheet_obj, end_row):
    for items in sorted(sheet_obj.merged_cell_ranges):
        print([items, end_row])
        _f, _e = str(items).split(":")
        if _f.find(COL_LIM) != -1:
            continue
        try:
            if int(_f[1:]) < ROW_LIM or int(_e[1:]) > end_row:
                continue
        except:
            pass
        sheet_obj.unmerge_cells(str(items))
        set_border(sheet_obj, str(items))
    return sheet_obj

def set_border(ws, cell_range):
    rows = ws[cell_range]

    rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):

        max_x = len(cells) - 1  # index of the last cell
        for pos_x, cell in enumerate(cells):
            cell.border =  Border(top=THIN, left=THIN, right=THIN, bottom=THIN)
